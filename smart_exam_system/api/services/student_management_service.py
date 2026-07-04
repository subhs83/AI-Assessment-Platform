from smart_exam_system.extensions import db
from smart_exam_system.models import StudentModel
from pathlib import Path
from werkzeug.exceptions import NotFound
from uuid import uuid4
from openpyxl import load_workbook
from smart_exam_system.api.services.react_student_service import create_student, normalize_text
from smart_exam_system.models.student import StudentRegistrationType


def validate_student_fields(
    first_name,
    student_class,
    roll_number,
):
    """
    Validate required student fields.

    Returns:
        dict: Field errors. Empty dict means valid.
    """

    errors = {}

    if not first_name or not first_name.strip():
        errors["first_name"] = "First Name is required."

    if not student_class or not student_class.strip():
        errors["student_class"] = "Class is required."

    if not roll_number or not roll_number.strip():
        errors["roll_number"] = "Roll Number is required."

    return errors


def is_duplicate_roll_number(
    school_id,
    student_class,
    roll_number,
    exclude_student_id=None,
):
    """
    Check if a roll number already exists within a class.
    """

    query = StudentModel.query.filter_by(
        school_id=school_id,
        student_class=student_class.strip(),
        roll_number=roll_number.strip(),
    )

    if exclude_student_id is not None:
        query = query.filter(
            StudentModel.id != exclude_student_id
        )

    return query.first() is not None


def get_students(school_id):

    students = (
        StudentModel.query
        .filter_by(school_id=school_id)
        .order_by(
            StudentModel.student_class.asc(),
            StudentModel.roll_number.asc(),
            StudentModel.first_name.asc()
        )
        .all()
    )

    return [
        {
            "student_uid": student.student_uid,
            "first_name": student.first_name,
            "last_name": student.last_name,
            "student_name": f"{student.first_name} {student.last_name}".strip(),
            "student_class": student.student_class,
            "roll_number": student.roll_number,
            "mobile": student.mobile,
            "student_registration_type": student.student_registration_type,
        }
        for student in students
    ]


def get_student_template():

    file_path = (
        Path(__file__).resolve().parents[2]
        / "static"
        / "downloads"
        / "student_import_template.xlsx"
    )


    if not file_path.exists():
        raise NotFound("Student template not found.")

    return file_path




def import_students(school_id, excel_file):
    if not excel_file:
        return False, "Please select an Excel file."

    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active

    except Exception:
        return False, "Invalid Excel file."

    expected_headers = [
        "First Name",
        "Last Name",
        "Class",
        "Roll Number",
        "Mobile",
    ]

    headers = [
        str(cell.value).strip() if cell.value else ""
        for cell in sheet[1]
    ]

    if headers != expected_headers:
        return (
            False,
            "Invalid template. Please download the latest student template.",
        )

    created = 0
    skipped_duplicates = 0
    invalid_rows = []

    existing_students = {
        (
            student.student_class.strip().lower(),
            student.roll_number.strip().lower(),
        )
        for student in StudentModel.query.filter_by(
            school_id=school_id
        ).all()
    }

    file_duplicates = set()

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):

        first_name = normalize_text(row[0], field="name")
        last_name = normalize_text(row[1], field="name")
        student_class = normalize_text(row[2], field="class")
        roll_number = normalize_text(row[3], field="roll_number")
        mobile = normalize_text(row[4], field="mobile")

        errors = validate_student_fields(
            first_name=first_name,
            student_class=student_class,
            roll_number=roll_number,
        )

        if errors:
            invalid_rows.append(
                {
                    "row": row_number,
                    "reason": next(iter(errors.values())),
                }
            )
            continue

        key = (
            student_class.lower(),
            roll_number.lower(),
        )

        if key in file_duplicates:
            skipped_duplicates += 1
            continue

        if key in existing_students:
            skipped_duplicates += 1
            continue

        file_duplicates.add(key)

        create_student(
            school_id=school_id,
            first_name=first_name,
            last_name=last_name,
            student_class=student_class,
            roll_number=roll_number,
            mobile=mobile,
            student_registration_type=StudentRegistrationType.VERIFIED,
        )

        created += 1

    db.session.commit()

    return (
        True,
        {
            "created": created,
            "skipped_duplicates": skipped_duplicates,
            "invalid_rows": invalid_rows,
        },
    )




def is_registration_allowed(exam, student):
    """
    Single source of truth for registration rules.
    Used by:
        - state
        - start
        - future submit/result flows
    """

    # OPEN MODE → always allowed
    if exam.registration_mode == "open":
        return True

    # VERIFIED MODE → student must exist
    if exam.registration_mode == "verified":
        if not student:
            return False

        return True

    # SAFETY DEFAULT (future modes)
    return False